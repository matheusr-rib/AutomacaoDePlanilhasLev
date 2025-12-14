from pathlib import Path
from typing import Dict, Any

import csv
from openpyxl import load_workbook

from padronizacao.servico_padronizacao import ServicoPadronizacao


def promover_padroes(caminho_validacao: Path):
    """
    Lê arquivo de validação humana (.csv OU .xlsx) e
    atualiza o cache JSON antes da execução principal.
    """
    servico = ServicoPadronizacao()
    atualizados = 0

    if caminho_validacao.suffix.lower() == ".csv":
        linhas = _ler_csv(caminho_validacao)

    elif caminho_validacao.suffix.lower() == ".xlsx":
        linhas = _ler_xlsx(caminho_validacao)

    else:
        raise ValueError(
            "Arquivo de validação deve ser .csv ou .xlsx"
        )

    for row in linhas:
        chave = (row.get("chave_cache") or "").strip()
        if not chave:
            continue

        aprovado = (row.get("aprovado") or "").strip().upper()
        if aprovado != "SIM":
            continue

        valor = {
            "produto_padronizado": (
                row.get("produto_corrigido")
                or row.get("ia_produto_padronizado")
            ),
            "convenio_padronizado": (
                row.get("convenio_corrigido")
                or row.get("ia_convenio_padronizado")
            ),
            "familia_produto": (
                row.get("familia_corrigida")
                or row.get("ia_familia_produto")
            ),
            "grupo_convenio": (
                row.get("grupo_corrigido")
                or row.get("ia_grupo_convenio")
            ),
        }

        servico.cache.set(chave, valor, overwrite=True)
        atualizados += 1

    if atualizados:
        servico.cache.salvar()

    print(f"Padrões atualizados no cache: {atualizados}")


# ==========================================================
# LEITORES
# ==========================================================

def _ler_csv(caminho: Path):
    with caminho.open("r", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f, delimiter=";")
        return list(reader)


def _ler_xlsx(caminho: Path):
    wb = load_workbook(caminho, data_only=True)
    ws = wb.active

    headers = []
    linhas = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c else "" for c in row]
            continue

        linha = {}
        for h, v in zip(headers, row):
            linha[h] = v

        linhas.append(linha)

    return linhas
