from typing import List, Dict, Any
from pathlib import Path

import pandas as pd


def escrever_excel(linhas: List[Dict[str, Any]], colunas: List[str], caminho_saida: Path) -> None:
    """
    Escreve um DataFrame em Excel, garantindo a ordem das colunas,
    e pinta de vermelho as linhas cuja padronização veio da IA.
    """
    # 1) monta DF só com colunas da planilha (não inclui chaves internas)
    df = pd.DataFrame(linhas, columns=colunas)

    caminho_saida.parent.mkdir(parents=True, exist_ok=True)

    # 2) escreve o excel
    df.to_excel(caminho_saida, index=False)

    # 3) pinta linhas IA (pós-processamento com openpyxl)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        wb = load_workbook(caminho_saida)
        ws = wb.active

        fill_vermelho = PatternFill(
            start_color="FFFFC7CE",  # vermelho claro (Excel style)
            end_color="FFFFC7CE",
            fill_type="solid",
        )

        # header é linha 1, dados começam na linha 2
        for i, linha in enumerate(linhas, start=2):
            origem = (linha.get("__ORIGEM_PADRONIZACAO") or "").upper()
            if origem in {"IA", "MANUAL"}:
                for col in range(1, len(colunas) + 1):
                    ws.cell(row=i, column=col).fill = fill_vermelho

        wb.save(caminho_saida)

    except Exception:
        # se openpyxl falhar por qualquer motivo, pelo menos o arquivo sai escrito
        return
